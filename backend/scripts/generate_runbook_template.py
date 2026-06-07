"""Generate a blank runbook Excel template.

Run from backend/:
    python -m scripts.generate_runbook_template [output_path]

Produces an .xlsx file with two sheets:
  - Runbook  (metadata key-value pairs)
  - Content  (flat hierarchy of phases/sections/tasks)

The file is ready to open in Excel / Google Sheets. Teams fill it in and
upload via POST /api/runbooks/upload or through the Directus admin upload page.
"""
from __future__ import annotations

import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed — run: pip install openpyxl")
    sys.exit(1)

# ── brand colours ─────────────────────────────────────────────────────────────
OCHRE    = "FF4900"
INK      = "0A0A0A"
PAPER2   = "F6F5F1"
RULE     = "E6E3DC"
WHITE    = "FFFFFF"
GOOD     = "22C55E"
SOFT     = "3F3F3F"

# ── style helpers ─────────────────────────────────────────────────────────────

def _header_style(cell, bg: str = OCHRE, fg: str = WHITE, bold: bool = True):
    cell.font      = Font(name="Calibri", bold=bold, color=fg, size=11)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

def _label_style(cell):
    cell.font      = Font(name="Calibri", bold=True, color=INK, size=10)
    cell.fill      = PatternFill("solid", fgColor=PAPER2)
    cell.alignment = Alignment(horizontal="left", vertical="top")

def _value_style(cell):
    cell.font      = Font(name="Calibri", color=SOFT, size=10)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

def _row_style(cell, bg: str = WHITE):
    cell.font      = Font(name="Calibri", color=SOFT, size=10)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

def _thin_border():
    s = Side(style="thin", color=RULE)
    return Border(left=s, right=s, top=s, bottom=s)

# ── Sheet 1: Runbook metadata ─────────────────────────────────────────────────

def _build_meta_sheet(wb: openpyxl.Workbook):
    ws = wb.active
    ws.title = "Runbook"
    ws.sheet_view.showGridLines = False

    # Title banner
    ws.merge_cells("A1:C1")
    banner = ws["A1"]
    banner.value = "DEPT® Runbook Template — Metadata"
    _header_style(banner)
    ws.row_dimensions[1].height = 30

    # Instructions
    ws.merge_cells("A2:C2")
    inst = ws["A2"]
    inst.value = (
        "Fill in each value in column B. Do not rename this sheet. "
        "Then fill the 'Content' sheet with your phases, sections and tasks."
    )
    inst.font      = Font(name="Calibri", italic=True, color=SOFT, size=9)
    inst.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    # Field rows
    fields = [
        ("Slug",        "banking-architect-greenfield",
         "Unique ID for this runbook. Use lowercase letters and hyphens only. "
         "E.g. banking-architect-greenfield"),
        ("Title",       "Banking Architect — Greenfield Engagement",
         "Full display title shown on the runbook page."),
        ("Role",        "architect",
         "architect | devops | developer | qa | pm | ba"),
        ("Domain",      "banking",
         "banking | ecommerce | manufacturing | healthcare | generic"),
        ("Type",        "greenfield",
         "greenfield | brownfield"),
        ("Description", "30-60-90 day engagement playbook for new banking builds on Adobe Experience Cloud.",
         "One or two sentences shown below the title on the reader page."),
        ("Status",      "draft",
         "draft (not visible to learners) | published (visible on /runbook index)"),
    ]

    for i, (field, example, hint) in enumerate(fields, start=3):
        ws.row_dimensions[i].height = 42
        lbl = ws.cell(row=i, column=1, value=field)
        val = ws.cell(row=i, column=2, value=example)
        tip = ws.cell(row=i, column=3, value=hint)
        _label_style(lbl)
        _value_style(val)
        tip.font      = Font(name="Calibri", italic=True, color="8A8780", size=9)
        tip.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        for cell in (lbl, val, tip):
            cell.border = _thin_border()

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 54


# ── Sheet 2: Content ──────────────────────────────────────────────────────────

CONTENT_COLS = [
    ("type",        14,  "phase | section | task | step | checklist | link"),
    ("title",       46,  "Main text for this row"),
    ("description", 52,  "Supporting detail (optional)"),
    ("owner",       18,  "Who does this? E.g. Architect, DevOps (task rows only)"),
    ("tools",       28,  "Comma-separated. E.g. AEM, Adobe IO, Jira (task rows only)"),
    ("timing",      18,  "When. E.g. Day 1, Week 2 (phase/section/task rows)"),
    ("url",         42,  "Full URL — used for link rows"),
    ("notes",       36,  "Optional freeform notes (task rows only)"),
]

TYPE_COLOURS = {
    "phase":     (OCHRE,  WHITE),
    "section":   (INK,    WHITE),
    "task":      (PAPER2, INK),
    "step":      (WHITE,  SOFT),
    "checklist": (WHITE,  SOFT),
    "link":      (WHITE,  SOFT),
}

SAMPLE_ROWS = [
    # (type, title, description, owner, tools, timing, url, notes)
    ("phase",     "Days 1–30: Onboarding",       "First month — environment, access, and discovery", "", "", "Days 1–30", "", ""),
    ("section",   "First Week",                  "Immediate priorities before first client meeting",  "", "", "Day 1–5",   "", ""),
    ("task",      "Set up development environment", "Clone repos, install SDKs, configure local AEM", "Architect", "AEM, Adobe IO, Git", "Day 1", "", "Use the standard DEPT® dev-env script from the repo wiki"),
    ("step",      "Clone the project repository", "", "", "", "", "https://github.com/dept/your-repo", ""),
    ("step",      "Run the local-setup script",   "Installs AEM 6.5 + dispatcher locally", "", "", "", "", ""),
    ("checklist", "VPN access confirmed",         "", "", "", "", "", ""),
    ("checklist", "Adobe IO console access granted", "", "", "", "", "", ""),
    ("link",      "DEPT® AEM starter kit",        "", "", "", "", "https://github.com/dept/aem-starter", ""),
    ("task",      "Introduce yourself to the client team", "", "Architect", "", "Day 1", "", ""),
    ("step",      "Send intro email with your working hours and primary contact", "", "", "", "", "", ""),
    ("checklist", "Client Slack / Teams channel joined", "", "", "", "", "", ""),
    ("section",   "Weeks 2–4",                   "Discovery and architecture baseline",               "", "", "Day 6–30",  "", ""),
    ("task",      "Run CODE-CODER discovery workshop", "Walk the client through all 14 nodes", "Architect", "Code-Coder Checklist", "Week 2", "", ""),
    ("step",      "Book a 3-hour working session with client architect + PO", "", "", "", "", "", ""),
    ("step",      "Walk through the checklist PDF with client", "", "", "", "", "https://internal.in.deptagency.com/anatomy/code-coder-checklist.html", ""),
    ("checklist", "All red/amber items have an owner and a date", "", "", "", "", "", ""),
    ("phase",     "Days 31–60: Integration",     "Build the first working vertical", "", "", "Days 31–60", "", ""),
    ("section",   "Architecture decisions",       "", "", "", "Week 5–6",  "", ""),
    ("task",      "Document ADRs for top 5 decisions", "", "Architect", "Confluence, ADR template", "Week 5", "", "One ADR per decision — see DEPT® ADR template"),
    ("step",      "List candidate decisions in the project wiki", "", "", "", "", "", ""),
    ("step",      "Run async review with engineering lead — 48 h comment window", "", "", "", "", "", ""),
    ("checklist", "ADRs reviewed and signed off by client architect", "", "", "", "", "", ""),
    ("phase",     "Days 61–90: Stabilisation",   "Handover readiness and ongoing ops", "", "", "Days 61–90", "", ""),
    ("section",   "Knowledge transfer",           "", "", "", "Week 10–12", "", ""),
    ("task",      "Record architecture walkthrough video", "", "Architect", "Loom, Confluence", "Week 11", "", ""),
    ("step",      "Record a 20-min Loom covering system topology, data flows, key ADRs", "", "", "", "", "", ""),
    ("checklist", "Video link added to project Confluence home", "", "", "", "", "", ""),
]


def _build_content_sheet(wb: openpyxl.Workbook):
    ws = wb.create_sheet("Content")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # Column header row
    for col_idx, (name, width, hint) in enumerate(CONTENT_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        _header_style(cell)
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 24

    # Hint row (row 2, greyed out, italic)
    for col_idx, (_, _, hint) in enumerate(CONTENT_COLS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=hint)
        cell.font      = Font(name="Calibri", italic=True, color="8A8780", size=9)
        cell.fill      = PatternFill("solid", fgColor=PAPER2)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border    = _thin_border()
    ws.row_dimensions[2].height = 20

    # Sample data rows
    col_names = [c[0] for c in CONTENT_COLS]
    for row_idx, sample in enumerate(SAMPLE_ROWS, start=3):
        row_type = sample[0]
        bg, fg = TYPE_COLOURS.get(row_type, (WHITE, SOFT))
        ws.row_dimensions[row_idx].height = 18
        for col_idx, value in enumerate(sample, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(name="Calibri", color=fg, size=10,
                                  bold=(row_type in ("phase", "section")))
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            cell.border    = _thin_border()

    # Add a few blank rows below for teams to fill in
    for row_idx in range(len(SAMPLE_ROWS) + 3, len(SAMPLE_ROWS) + 20):
        ws.row_dimensions[row_idx].height = 18
        for col_idx in range(1, len(CONTENT_COLS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx, value="")
            cell.fill   = PatternFill("solid", fgColor=WHITE)
            cell.border = _thin_border()


# ── Legend sheet ──────────────────────────────────────────────────────────────

def _build_legend_sheet(wb: openpyxl.Workbook):
    ws = wb.create_sheet("Legend")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    banner = ws["A1"]
    banner.value = "Row type reference"
    _header_style(banner)
    ws.row_dimensions[1].height = 28

    rows = [
        ("phase",     "Top-level block (e.g. Days 1–30). Sets timing for the block.",
         "Required for each major phase. Appears as a large section header."),
        ("section",   "Sub-group within a phase (e.g. First Week, Architecture).",
         "Optional — use when a phase has multiple distinct sub-areas."),
        ("task",      "Concrete thing to do. The main unit of the runbook.",
         "Fill owner, tools, and timing here. Steps/checklist go below."),
        ("step",      "Numbered action inside a task (title = the step text).",
         "url column can hold a link for this step. description adds context."),
        ("checklist", "Checkbox item inside a task (title = the item text).",
         "Keep items short. No url or owner needed."),
        ("link",      "Reference link inside a task (title = link label, url = href).",
         "Use for docs, tools, templates. Appears as a clickable link."),
    ]

    ws.cell(row=2, column=1, value="type").font = Font(bold=True, color=INK)
    ws.cell(row=2, column=2, value="what it does").font = Font(bold=True, color=INK)
    ws.cell(row=2, column=3, value="tips").font = Font(bold=True, color=INK)

    for i, (rtype, what, tip) in enumerate(rows, start=3):
        bg, fg = TYPE_COLOURS.get(rtype, (WHITE, SOFT))
        for col, val in [(1, rtype), (2, what), (3, tip)]:
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(name="Calibri", color=fg, size=10,
                                  bold=(rtype in ("phase", "section")))
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border    = _thin_border()
        ws.row_dimensions[i].height = 36

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 62


# ── entry point ───────────────────────────────────────────────────────────────

def generate(output_path: Path):
    wb = openpyxl.Workbook()
    _build_meta_sheet(wb)
    _build_content_sheet(wb)
    _build_legend_sheet(wb)
    wb.save(output_path)
    print(f"Template saved → {output_path}")
    print()
    print("Next steps:")
    print("  1. Open the file in Excel or Google Sheets.")
    print("  2. Edit the 'Runbook' sheet — especially Slug, Title, Role, Domain.")
    print("  3. Fill the 'Content' sheet — replace sample rows with your runbook content.")
    print("  4. Save as .xlsx and upload via:")
    print("     POST /api/runbooks/upload   (multipart/form-data, field: file)")
    print("     Or use the admin page at /app/#/admin/runbooks")
    print()
    print("Auth: the upload endpoint requires the content_author role.")
    print("Ask a platform_admin to grant it: POST /api/admin/roles {email, role_key: content_author}")


if __name__ == "__main__":
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("data/runbook-template.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    generate(out)
