"""Excel → Runbook JSON parser.

Template structure
──────────────────
Sheet 1 — "Runbook"  (metadata, key-value pairs)
  Row format:  | Field Name | Value |
  Fields: Slug, Title, Role, Domain, Type, Description, Status

Sheet 2 — "Content"  (flat hierarchy driven by the `type` column)
  Header row: type | title | description | owner | tools | timing | url | notes

  type values (case-insensitive):
    phase      — starts a new phase block
    section    — starts a new section within the current phase
    task       — a task within the current section
    step       — numbered step inside the current task  (title = step text)
    checklist  — checkbox item inside the current task  (title = item text)
    link       — reference link inside the current task (title = label, url = href)

  Empty rows, rows with no type, and comment rows (type starts with #) are skipped.

Validation errors raise ValueError with a human-readable message so the
upload endpoint can return a 422 with context.
"""
from __future__ import annotations

import re
from io import BytesIO
from typing import Any, Dict, List, Optional

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise ImportError(
        "openpyxl is required for Excel parsing — install it with: pip install openpyxl"
    ) from exc

from app.modules.runbooks.schemas import (
    RunbookCreate,
    RunbookLink,
    RunbookPhase,
    RunbookSection,
    RunbookTask,
)

# ── helpers ─────────────────────────────────────────────────────────────────

def _cell(v: Any) -> str:
    """Normalise a cell value to a stripped string, empty string if None."""
    return str(v).strip() if v is not None else ""


def _tools(raw: str) -> List[str]:
    """Split a comma-separated tools cell into a list, skipping empties."""
    return [t.strip() for t in raw.split(",") if t.strip()]


# ── metadata sheet ───────────────────────────────────────────────────────────

def _parse_meta(wb: "openpyxl.Workbook") -> Dict[str, str]:
    """Read the 'Runbook' sheet as field→value pairs (case-insensitive keys)."""
    if "Runbook" not in wb.sheetnames:
        raise ValueError(
            "Missing 'Runbook' sheet. The workbook must have a sheet named exactly 'Runbook' "
            "with rows like: | Slug | banking-architect-greenfield |"
        )
    ws = wb["Runbook"]
    meta: Dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        key = _cell(row[0] if row else None).lower().replace(" ", "_")
        val = _cell(row[1] if len(row) > 1 else None)
        if key and val:
            meta[key] = val
    return meta


# ── content sheet ────────────────────────────────────────────────────────────

def _parse_content(wb: "openpyxl.Workbook") -> List[RunbookPhase]:
    """Read the 'Content' sheet and return the phases tree."""
    if "Content" not in wb.sheetnames:
        raise ValueError(
            "Missing 'Content' sheet. The workbook must have a sheet named exactly 'Content' "
            "with a header row: type | title | description | owner | tools | timing | url | notes"
        )
    ws = wb["Content"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Build column index from the header row
    header = [_cell(c).lower() for c in rows[0]]
    required = {"type", "title"}
    missing = required - set(header)
    if missing:
        raise ValueError(
            f"Content sheet header is missing required columns: {sorted(missing)}. "
            f"Found: {header}"
        )

    def col(row_vals: tuple, name: str, default: str = "") -> str:
        try:
            idx = header.index(name)
            return _cell(row_vals[idx]) if idx < len(row_vals) else default
        except ValueError:
            return default

    phases: List[RunbookPhase] = []
    cur_phase: Optional[RunbookPhase] = None
    cur_section: Optional[RunbookSection] = None
    cur_task: Optional[RunbookTask] = None

    for row_num, row in enumerate(rows[1:], start=2):
        row_type = col(row, "type").lower()

        # Skip empty, blank, or comment rows
        if not row_type or row_type.startswith("#"):
            continue

        title   = col(row, "title")
        desc    = col(row, "description") or None
        owner   = col(row, "owner") or None
        tools   = _tools(col(row, "tools"))
        timing  = col(row, "timing") or None
        url     = col(row, "url") or None
        notes   = col(row, "notes") or None

        if row_type == "phase":
            cur_phase   = RunbookPhase(title=title, description=desc, timing=timing)
            cur_section = None
            cur_task    = None
            phases.append(cur_phase)

        elif row_type == "section":
            if cur_phase is None:
                # Implicit top-level phase for section without a preceding phase row
                cur_phase = RunbookPhase(title="General")
                phases.append(cur_phase)
            cur_section = RunbookSection(title=title, description=desc, timing=timing)
            cur_task    = None
            cur_phase.sections.append(cur_section)

        elif row_type == "task":
            if cur_section is None:
                if cur_phase is None:
                    cur_phase = RunbookPhase(title="General")
                    phases.append(cur_phase)
                cur_section = RunbookSection(title="General")
                cur_phase.sections.append(cur_section)
            cur_task = RunbookTask(
                title=title,
                description=desc,
                owner=owner,
                tools=tools,
                timing=timing,
                notes=notes,
            )
            cur_section.tasks.append(cur_task)

        elif row_type in ("step", "checklist", "link"):
            if cur_task is None:
                # Silently skip orphan step/checklist/link rows (bad data — don't abort)
                continue
            if row_type == "step":
                text = title + (f" — {desc}" if desc else "")
                cur_task.steps.append(text)
            elif row_type == "checklist":
                cur_task.checklist.append(title)
            elif row_type == "link":
                if url:
                    cur_task.links.append(RunbookLink(label=title or url, url=url))

        else:
            # Unknown type — warn in logs but do not abort
            import logging
            logging.getLogger("runbooks.parser").warning(
                "Row %d: unknown type %r — skipped", row_num, row_type
            )

    return phases


# ── public entry point ───────────────────────────────────────────────────────

def parse_excel(file_bytes: bytes) -> RunbookCreate:
    """Parse an Excel workbook (bytes) into a validated RunbookCreate object.

    Raises ValueError with a descriptive message on any structural problem.
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Could not open workbook: {exc}") from exc

    meta   = _parse_meta(wb)
    phases = _parse_content(wb)

    slug = meta.get("slug", "")
    if not slug:
        raise ValueError(
            "The 'Runbook' metadata sheet must have a 'Slug' row with a value, "
            "e.g.: | Slug | banking-architect-greenfield |"
        )

    title = meta.get("title", slug.replace("-", " ").title())

    return RunbookCreate(
        slug=slug,
        title=title,
        role=meta.get("role", "architect"),
        domain=meta.get("domain", "generic"),
        type=meta.get("type", "greenfield"),
        description=meta.get("description") or None,
        status=meta.get("status", "draft"),
        phases=phases,
    )
